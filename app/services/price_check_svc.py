"""Price-check service: parse Guardian.xlsx → compare weeks → generate DFI output xlsx."""
from __future__ import annotations

import base64
import io
import re
from collections import defaultdict
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ── Master SKU list (fixed 107 SKUs) ──────────────────────────────────────────
# (stt, article, barcode, description, product_name)
_MASTER_SKUS: list[tuple[int, str, str, str, str]] = [
    (1,   "VN-GD-3021691", "8809647390015", "SOME BY MI CLEAR SPOT PATCH 18P",                                          "01.SOME BY MI - Miếng Dán Some By Mi Giảm Mụn Thần Kỳ Cho Mọi Loại Da Spot Patch 18 M"),
    (2,   "VN-GD-3019062", "4719745504365", "DERMA ANGEL ACNE PATCH FOR DAY 12PCS",                                    "02.DERMA ANGEL - Miếng Dán Mụn Dermaangel Dùng Ban Ngày 12 Miếng"),
    (3,   "VN-GD-3029141", "3337875863377", "LRP EFFACLAR DUO+ M MOISTURISER FOR OILY ACNE SKIN 40ML",                 "03.LA ROCHE-POSAY - Kem Dưỡng Giảm Mụn La Roche Posay Effaclar Duo+M 40ml"),
    (4,   "VN-GD-3019079", "8934868167112", "LIFEBUOY SHOWER GEL MATCHA MELON 850G/12",                                "04.LIFEBUOY - Sữa Tắm Lifebuoy Matcha & Khổ Qua Detox Da Khỏi Vi Khuẩn Gây Mụn 800g"),
    (5,   "VN-GD-3021168", "8936217700032", "COCOON DAK LAK COFFEE BODY POLISH 200ML/",                                "05.COCOON - Tẩy Da Chết Cocoon Dak Lak Coffee Body Polish Từ Cà Phê Đak Lak 200ml"),
    (6,   "VN-GD-3027232", "6902088325454", "DOVE TAY TBC LUU DO 6X(298G)",                                            "06.DOVE - Tẩy Tế Bào Chết Dưỡng Ẩm Toàn Thân Dove Hương Lựu Đỏ 298g"),
    (7,   "VN-GD-3021518", "4513574027077", "REIHAKU HATOMUGI SG MOISTURIZING 800ML",                                  "07.REIHAKU HATOMUGI - Sữa Tắm Reihaku Hatomugi Body Soap Làm Sáng Da 800ml"),
    (8,   "VN-GD-3008701", "9556006060438", "J&J ALMONDS SHOWER GELPH5.5 750ML",                                       "08.JOHNSON & JOHNSON - Sữa Tắm Johnson's PH 5.5 Hạnh Nhân 750ml"),
    (9,   "VN-GD-3030585", "8936217700933", "COCOON PALMYRA PALM SUGAR BODY POLISH 200ML",                             "09.COCOON - Tẩy Da Chết Cocoon Đường Thốt Nốt An Giang 200ml"),
    (10,  "VN-GD-3030790", "12044051283",   "OLD SPICE TIMBER WITH SANDALWOOD BODY WASH 532ML",                        "10.OLD SPICE - Sữa Tắm Old Spice Timber Sandal Wood 532ml"),
    (11,  "VN-GD-3013498", "8938531758225", "JOMI COTTON PADS 120P",                                                   "11.JOMI - Bông Tẩy Trang Jomi 120 miếng/Túi"),
    (12,  "VN-GD-3030773", "6977057251034", "MLEN DIARY SELF-ADHESIVE EYELASHES - SUNFLOWER EYELASHES",                "12.MLEN DIARY - Mi Giả Tự Dính Mlen Diary Mi Hướng Dương"),
    (13,  "VN-GD-3026634", "6902395975908", "MAYBELLINE MASCARA LASH SENSATIONAL SKY HIGH 6ML",                        "13.MAYBELLINE - Mascara Không Lem Trôi Maybelline Lash Sensational Sky High Waterproof 6ml.#Very Black"),
    (14,  "VN-GD-3025701", "6902395828426", "MAYBELLINE HYPER SHARP EXTREME LINER ULTRA BLACK 0.4G",                   "14.MAYBELLINE - Bút Kẻ Mắt Nước Maybelline New York Hyper Sharp Extreme Liner 0.4g .#BK-1 Ultra Black"),
    (15,  "VN-GD-3003399", "6923492584206", "MBL MASCARA HYPERCURL BLACK 9.2ML",                                       "15.MAYBELLINE - Mascara Maybelline New York Cong Mi Hypercurl 9.2ml - Đen"),
    (16,  "VN-GD-3024042", "8858842084090", "BROWIT MY EV.DAY MASCARA#ENDLESS NIGHT",                                  "16.BROWIT - Mascara Browit By Nongchat My Everyday Mascara 5.5g .#Endless Night"),
    (17,  "VN-GD-3021213", "8859762800180", "LEMONADE D.EYEBROW WIGINATURALBROWN2.75G",                                "17.LEMONADE - Kẻ mày Lemonade Dual Eyebrow 0.25g + 2.5g #02 Natural Brown"),
    (18,  "VN-GD-3021212", "8859762800173", "LEMONADE D.EYEBROW WIGI DARK BROWN 2.75G",                                "18.LEMONADE - Kẻ mày Lemonade Dual Eyebrow 0.25g + 2.5g #01 Dark Brown"),
    (19,  "VN-GD-3029590", "6975425384681", "LEMONADE AESTHETIC EYESHADOW PALETTE VER2",                               "19.LEMONADE - Bảng Phấn Mắt 16 Màu Lemonade Aesthetic Eyeshadow Palette 20.8g Version 2"),
    (20,  "VN-GD-3030371", "6902395970040", "SUPERSTAY LUMI MATTE 111 AS X",                                           "20.MAYBELLINE - Kem Nền Lâu Trôi Maybelline Super Stay Up To 30H Lumi-Matte Foundation 35ml .#111"),
    (21,  "VN-GD-3030370", "6902395970033", "SUPERSTAY LUMI MATTE 110 AS X",                                           "21.MAYBELLINE - Kem Nền Lâu Trôi Maybelline Super Stay Up To 30H Lumi-Matte Foundation 35ml .#110"),
    (22,  "VN-GD-3030350", "6976436772573", "JUDYDOLL HIGHLIGHT & CONTOUR02",                                          "22.JUDYDOLL - Phấn Bắt Sáng & Tạo Khối Judydoll Highlight & Contour 9g .#02"),
    (23,  "VN-GD-3026770", "6902395828327", "MAYBELLINE FIT ME BLUSH #30 FIERCE 4.5G",                                 "23.MAYBELLINE - Phấn Má Hồng Maybelline Fit Me Mono Blush Mịn Mướt 16h 4.5g .# 30 Fierce"),
    (24,  "VN-GD-3031475", "6977304930958", "COLORKEY AIRY MATTE SETTING SPRAY FOR OILY SKIN - VERSION 2",             "24.COLORKEY - Xịt Khóa Lớp Trang Điểm Colorkey Airy Soft Matte Makeup Setting Spray V2 100ml"),
    (25,  "VN-GD-3030349", "6976436772566", "JUDYDOLL HIGHLIGHT & CONTOUR01",                                          "25.JUDYDOLL - Phấn Bắt Sáng & Tạo Khối Judydoll Highlight & Contour 9g .#01"),
    (26,  "VN-GD-3030421", "6926799657529", "ZA TW PROTECTOR 02PURPLE RENEWAL",                                        "26.ZA - Kem Lót Nâng Tông Chống Nắng ZA True White Glowing Protector SPF36 PA++ 35g .#02 Purple"),
    (27,  "VN-GD-3031593", "3600531668440", "MBL SUNKISSER LIQUID BLUSH 06 4.7ML",                                     "27.MAYBELLINE - Má Hồng Dạng Kem Maybelline Sunkisser Blush 4.7g .#06 City Sizzle Cánh Hồng Khô"),
    (28,  "VN-GD-3028095", "8809664983931", "3CE SHIMMER MAKEUP FIXER 100ML",                                          "28.3CE - Xịt Khoáng Nhũ 3CE Cố Định Lớp Trang Điểm Và Tạo Sự Căng Bóng Shimmer Makeup Fixer 95ml"),
    (29,  "VN-GD-3026891", "8809721243251", "BOM COVER FLEX CUSHION SPF50+ PA+++ #21N IVORY 15G",                      "29.BOM - Phấn Nước Đa Năng BOM Che Mọi Khuyết Điểm Cover Flex Cushion SPF50+ 15g .#21N Ivory Màu Trắng Ngà"),
    (30,  "VN-GD-3016898", "41554247701",   "MBL FITME CONCEALER 15 FAIR 6.8ML",                                       "30.MAYBELLINE - Kem Che Khuyết Điểm Maybelline New York Fit Me 6,8ml - Fair"),
    (31,  "VN-GD-3030976", "6975326789936", "COLORKEY LIGHT WEIGHT POLISH PRIMER 01",                                  "31.COLORKEY - Kem Lót Colorkey Light Weight Polish Primer 30g .#01"),
    (32,  "VN-GD-3026714", "6974614885015", "PERFECT DIARY WEIGHTLESS SOFT-VELVET BLURRING LOOSE POWDER #01 7G",       "32.THE PERFECT DIARY - Phấn Phủ Kiềm Dầu Perfect Diary Weightless Soft-Velvet Blurring Loose Powder 7g .#01"),
    (33,  "VN-GD-3025981", "8809450450371", "LEMONADE CUSHION SUPERMATTE A01 15G",                                     "33.LEMONADE - Phấn Nước Kiềm Dầu Lemonade SuperMatte SPF50+ PA+++ 15g .#A01 Light"),
    (34,  "VN-GD-3030401", "6940412453442", "CARSLAN SOFT FOCUS MAKEUP POWDER 01T PURPLE 8G",                          "34.CARSLAN - Phấn Phủ Dạng Bột Carslan Soft Focus Make-Up Powder 8g .#01 (Purple)"),
    (35,  "VN-GD-3016983", "305210231597",  "VASELINE LIP BALM ROSY 7G/32",                                            "35.VASELINE - Sáp Dưỡng Môi Vaseline Hồng Xinh 7g"),
    (36,  "VN-GD-3024799", "4511413503997", "DHC LIP CREAM 1.5G",                                                      "36.DHC - Son dưỡng DHC Lip Cream"),
    (37,  "VN-GD-3024747", "4550516703552", "SENKA PERFECT AQUA LIP ESSENCE 10G",                                      "37.SENKA - Tinh Chất Dưỡng Môi Senka Perfect Aqua Lip Essence Giúp Cấp Ẩm 10g"),
    (38,  "VN-GD-3025144", "8809584962412", "CARENEL POMEGRANATE LIP NIGHT MASK 5G",                                   "38.CARE:NEL - Mặt Nạ Ngủ Môi Care:Nel Pomegranate Lip Night Mask Cấp Ẩm Ngừa Nứt Nẻ Hương Lựu 5g"),
    (39,  "VN-GD-3019401", "8801619047545", "VASELINE LIP BALM ROSY LIPS 4.8G",                                        "39.VASELINE - Son Vaseline_Dưỡng Môi Hồng Xinh 4.8g"),
    (40,  "VN-GD-3028301", "6902395832676", "MBL SUPERSTAY VINYL INK LIPSTICK_10 EXTRA",                               "40.MAYBELLINE - Son Kem Bóng Maybelline Super Stay Vinyl Ink 4.2ml .#10 Lippy"),
    (41,  "VN-GD-3031579", "6902395972976", "3CE CASHMEREHUG LIPSTICK KN 07 ENKR",                                     "41.3CE - Son 3CE Cashmere Hug Lipstick 3.5g .#07 Knit Hồng Khô"),
    (42,  "VN-GD-3026348", "8809664983511", "3CE BLUR WATER TINT #SEPIA 4.6 G",                                        "42.3CE - Son Kem Lì 3CE Blur Water Tint 4.6g .# Sepia"),
    (43,  "VN-GD-3021547", "12044040898",   "OLD SPICE TIMBER DEO STICK 73G",                                          "43.OLD SPICE - Sáp Khử Mùi Old Spice Timber Hương Gỗ Đàn Hương 73g"),
    (44,  "VN-GD-3021540", "12044038918",   "OLD SPICE PURE SPORT DEO STICK 85G",                                      "44.OLD SPICE - Sáp Khử Mùi Old Spice Pure Sport High Endurance Dành Cho Nam Giới Chơi Thể Thao Và Vận Động 85g"),
    (45,  "VN-GD-3019084", "47400097728",   "GILLETTE GEL DEO COOL WAVE 107GR/12",                                     "45.GILLETTE - Gel Khử Mùi Gillette Cool Wave 107g"),
    (46,  "VN-GD-3027124", "3614819997108", "ETIAXIL DÉTRANSPIRANT TRAITEMENT TRANSPIRATION EXCESSIVE AISSELLES PEAUX SENSIBLES 15ML", "46.ETIAXIL - Lăn Khử Mùi Etiaxil Dành Cho Da Nhạy Cảm Aisselles Sensibles 15ml"),
    (47,  "VN-GD-3018205", "4902430734998", "OLAY REGENERIST ANTIAGING NIGHT CREAM50G",                                "47.OLAY - Kem Vi Dưỡng Ban Đêm Olay Regenerist 50g"),
    (48,  "VN-GD-3026450", "3337875597357", "CERAVE FOAMING CLEANSER 473ML",                                           "48.CERAVE - Sữa Rửa Mặt CeraVe Foaming Cleanser Giúp Làm Sạch Sâu Dành Cho Da Dầu 473ml"),
    (49,  "VN-GD-3004658", "8935006537675", "HADALABO ADVANCED NOURISH CLEANSER 80G",                                  "49.HADALABO - Kem Rửa Mặt Dưỡng Ẩm Tối Ưu Hada Labo Advanced Nourish 80g"),
    (50,  "VN-GD-3029073", "6975326789318", "COLORKEY HYDRATING FACIAL MASK 25ML",                                     "50.COLORKEY - Mặt Nạ Dưỡng Ẩm Colorkey Hydrating Facial Mask 25ml"),
    (51,  "VN-GD-3024355", "8936217700490", "COCOON HUNG YEN TUMERIC FACE MASK 30ML",                                  "51.COCOON - Mặt Nạ Nghệ Hưng Yên Cocoon Turmeric Face Mask Giúp Da Rạng Rỡ & Mịn Màng 30ml"),
    (52,  "VN-GD-3018713", "8938553765027", "BNBG VITA HYDRATING JELLY MASK 30ML",                                     "52.BNBG - Mặt Nạ BNBG Vita Genic Hydrating Jelly Mask Dưỡng Ẩm 30ml"),
    (53,  "VN-GD-3026285", "8809486363799", "BANOBAGI STEM CELL VITAMIN MASK WHITENING AND MOISTURE 30G",              "53.BANOBAGI - Mặt Nạ Banobagi Stem Cell Vitamin Mask Cấp Ẩm, Sáng Da Whitening & Moisture 30g"),
    (54,  "VN-GD-3030552", "8809544631372", "WONJIN NOURISING SUPPLEMENT MASK 30G",                                    "54.WONJIN - Mặt Nạ Wonjin Phục Hồi Da Effect Nourishing Supplement Concentrated Essence Mask 30g"),
    (55,  "VN-GD-3004659", "8935006537729", "HADALABO ADVANCED NOURISH CREAM 50G/20",                                  "55.HADALABO - Kem Dưỡng Ẩm Tối Ưu Hada Labo Advanced Nourish 50g"),
    (56,  "VN-GD-3025119", "8994993016587", "L'OREAL PARIS GLYCOLIC-BRIGHT INSTANT GLOWING SERUM VỚI 1.0% GLYCOLIC ACID", "56.L'ORÉAL - Tinh Chất L'Oreal Paris Dưỡng Sáng Da Glycolic-Bright Anti-Dark Spot Brightening Serum 30ml"),
    (57,  "VN-GD-3021765", "8809429958709", "9WISHES HYDRA SKIN AMPULE SERUM 30ML",                                    "57.9WISHES - Tinh Chất 9 Wishes Hydra Ampule II 30ml"),
    (58,  "VN-GD-3032054", "4987176287359", "OLAY SUPER SERUM 5 IN 1 30ML",                                            "58.OLAY - Tinh Chất Olay Super Serum 5 In 1 30ml"),
    (59,  "VN-GD-3021000", "8809115029119", "KLAIRS SUPPLE PREP UNSCENTED TONER 180ML",                                "59.KLAIRS - Nước Hoa Hồng Không Mùi Dear Klairs Dưỡng Ẩm Da Và Làm Mềm Da Supple Preparation Unscented 180ml"),
    (60,  "VN-GD-3030168", "8936217700889", "COCOON HAU GIANG LOTUS SOOTHING TONER 310ML",                             "60.COCOON - Nước Cân Bằng Cocoon Nước Sen Hậu Giang 310ml"),
    (61,  "VN-GD-3020942", "4902430733830", "OLAY WHITE RADIANT NIGHT CREAM 50G",                                      "61.OLAY - Kem Dưỡng Ban Đêm Olay Luminous Light Perfecting Night Cream 50g"),
    (62,  "VN-GD-3017388", "8936123410131", "LACTACYD FEMININE ODOR FRESH 150ML/24",                                   "62.LACTACYD - Dung Dịch Vệ Sinh Phụ Nữ Lactacyd Odor Fresh 150ml"),
    (63,  "VN-GD-3022696", "8934755014161", "DIANA SENSI DAILY ANTI-BAC LINER 40M",                                    "63.DIANA - Băng Vệ Sinh Diana Sensi Hằng Ngày Kháng Khuẩn 40 Miếng"),
    (64,  "VN-GD-3029381", "8934755010453", "DIANA SUPER NIGHT NAPKIN SHORT M-L 5M",                                   "64.DIANA - Băng Vệ Sinh Dạng Quần Diana Sensi Chống Tràn Size M-L 5 Chiếc/Gói"),
    (65,  "VN-GD-3020242", "8934755010330", "DIANA ULTRA WING NAPKIN 20M/24",                                          "65.DIANA - Băng Vệ Sinh Diana Siêu Thấm Siêu Mỏng Cánh Gói Lớn 20 Miếng"),
    (66,  "VN-GD-3019485", "8851818564985", "LAURIER FRE&FREE UT.SLIM W NAPKIN 20M/45",                                "66.LAURIER - Băng Vệ Sinh Laurier Fresh & Free Siêu Mỏng Có Cánh 20 Miếng"),
    (67,  "VN-GD-3020240", "8934755014321", "DIANA SENSI COOL FRESH DAILY 40M/48",                                     "67.DIANA - Băng Vệ Sinh Diana Sensi Cool Fresh Hằng Ngày 40 Miếng"),
    (68,  "VN-GD-3017768", "8851818070806", "LAURIER SUPPER SLIMGUARD 22.5CM 20M",                                     "68.LAURIER - Băng Vệ Sinh Laurier Siêu Mỏng Bảo Vệ 22Cm 20 Miếng"),
    (69,  "VN-GD-3020241", "8934755010439", "DIANA SENSI COOL FRESH UT WING 20M/24",                                   "69.DIANA - Băng Vệ Sinh Diana Sensi Cool Fresh Siêu Mỏng Có Cánh 23cm 20 Miếng"),
    (70,  "VN-GD-3021259", "8936123410254", "LACTACYD FEMININE ODOR FRESH 250ML/24",                                   "70.LACTACYD - Dung Dịch Vệ Sinh Phụ Nữ Lactacyd Odor Fresh 250ml"),
    (71,  "VN-GD-3001812", "8935006533363", "SELSUN ANTI DANDRUFF SHAMPOO 100ML",                                      "71.SELSUN - Dầu Gội Selsun Anti-Dandruff Dành cho Tóc Gàu 100ml"),
    (72,  "VN-GD-3026731", "4901872837144", "FINO PREMIUM TOUCH 230 G",                                                "72.FINO - Kem Ủ Tóc Cao Cấp Fino Phục Hồi Hư Tổn Premium Touch 230g"),
    (73,  "VN-GD-3027226", "8850006930526", "PALMOLIVE SHAMPOO INTENSIVE MOISTURE 600ML",                              "73.PALMOLIVE - Dầu Gội và Xả Palmolive Naturals Intensive Moisture Coconut Cream 600ml"),
    (74,  "VN-GD-3028208", "8936217700094", "COCOON SH PAMELO 500ML",                                                  "74.COCOON - Dầu Gội Bưởi Cocoon Giảm Gãy Rụng và Làm Mềm Tóc Pomelo Shampoo 500ml"),
    (75,  "VN-GD-3027529", "8936217700131", "COCOON POMELO HAIR TONIC 140ML",                                          "75.COCOON - Nước Dưỡng Tóc Cocoon Tinh Dầu Bưởi Pomelo Hair Tonic 140ml"),
    (76,  "VN-GD-3025478", "6923700914047", "L'OREAL PARIS EXTRAORDINARY OIL SHAMPOO 440ML",                          "76.L'ORÉAL - Dầu Gội L'oreal Paris Extraordinary Oil Dưỡng Tóc Suôn Mượt 440ml"),
    (77,  "VN-GD-3019714", "4992944111576", "LOREAL EL. HAIR EXTRAORDINARY OIL 100ML",                                 "77.L'ORÉAL - Dầu Dưỡng L'Oreal Paris Dưỡng Tóc Suôn Mượt Bồng Bềnh 100ml"),
    (78,  "VN-GD-3001392", "8935030218243", "DOUBLE RICH HAIR NUTRITION 250ML",                                        "78.DOUBLE RICH - Nước Dưỡng Tóc Double Rich Phục Hồi (Hồng) 250ml"),
    (79,  "VN-GD-3019096", "4902430624053", "PANTENE COND 3MM DAMAGE CARE 300ML/12",                                   "79.PANTENE - Dầu Xả Pantene Collagen Repair Intensive Serum Conditioner 300ml"),
    (80,  "VN-GD-3029093", "2000030290936", "BUNDLE DRY SH DAZZLING VOLUME 200ML",                                     "80.GIRLZ ONLY - Dầu Gội Khô Girlz Only Dazzling Volume Làm Phồng Tóc 200ml"),
    (81,  "VN-GD-3022823", "22796916709",   "OGX SP THICK&FULL BIOTIN COLLAGEN 385ML",                                 "81.OGX - Dầu Gội OGX Thick & Full Bổ Sung Biotin Collagen Cho Tóc 385ml"),
    (82,  "VN-GD-3016843", "6928820030226", "LOREAL MICELLAR REFRESHING 3IN1 400ML",                                   "82.L'ORÉAL - Tẩy Trang L'Oreal Paris Skincare Make Up Remover Micellar Refreshing Tươi Mát 400ml"),
    (83,  "VN-GD-3024783", "8936217700223", "COCOON WINTER MELON MICELLAR WATER 500ML",                                "83.COCOON - Nước Tẩy Trang Cocoon Winter Melon Bí Đao Làm Sạch Da Và Giảm Dầu 500ml"),
    (84,  "VN-GD-3017868", "3701129814420", "BIODERMA SENSIBIO H2OMICELLARWATER 500ML",                                "84.BIODERMA - Nước Tẩy Trang Dành Cho Da Nhạy Cảm Bioderma Sensibio H20 500ml"),
    (85,  "VN-GD-3024480", "8994993019670", "GARNIER MICELLAR OILY ACNE SKIN 400ML",                                   "85.GARNIER - Nước Tẩy Trang Garnier Micellar Cleansing Water Salicylic BHA 400ml"),
    (86,  "VN-GD-3001439", "4005808355754", "NIVEA BODY LOTION WHITE NIGHT 350ML/12",                                  "86.NIVEA - Sữa Dưỡng Thể Dưỡng Trắng Da Ban Đêm Nivea 350ml"),
    (87,  "VN-GD-3011290", "8934839121037", "CLOSE UP TP DIAMOND ATTRACTION 100G/36",                                  "87.CLOSE UP - Kem Đánh Răng Dạng Gel Closeup White Attraction Diamond Trắng Sáng 100g"),
    (88,  "VN-GD-3022652", "8850006932827", "COLGATE TOOTHBRUSH CUSHION CLEAN PACK2",                                  "88.COLGATE - Bộ Bàn Chải Đánh Răng Colgate Cushion Clean Twin Mềm Mại 2 Cây"),
    (89,  "VN-GD-3029825", "6920354836930", "COLGATE OPTIC WHITE PURPLE TOOTHPASTE 100GVN",                            "89.COLGATE - Kem Đánh Răng Colgate Optic White Purple 100g"),
    (90,  "VN-GD-3019467", "6920354814471", "COLGATE TP OPTICWHITE PLUS SHINE 100G/48",                               "90.COLGATE - Kem Đánh Răng Colgate Optic White Làm Trắng & Sáng Bổ Sung 100g"),
    (91,  "VN-GD-3029614", "8004395111718", "MARVIS TOOTHPASTE WHITENING MINT 85ML",                                   "91.MARVIS - Kem Đánh Răng Marvis Whitening Mint 85ml"),
    (92,  "VN-GD-3024755", "5011309895513", "EUCRYL WHITERNING TOOTHPASTE 62G",                                        "92.EUCRYL - Kem Đánh Răng Tẩy Trắng Eucryl Toothpaste 62g"),
    (93,  "VN-GD-3018335", "8850090400424", "SENSODYNE TP GENTLE WHITENING 160G/48",                                   "93.SENSODYNE - Kem Đánh Răng Trắng Sáng Sensodyne Gentle Whitening 160g"),
    (94,  "VN-GD-3024754", "5011309895612", "EUCRYL WHITERNING TOOTHPOWDER 50G",                                       "94.EUCRYL - Bột Đánh Răng Tẩy Trắng Eucryl Toothpowder 50g"),
    (95,  "VN-GD-3029615", "8004395111817", "MARVIS TOOTHPASTE SMOKERS WHITENING MINT 85ML",                           "95.MARVIS - Kem Đánh Răng Marvis Smokers Whitening Mint 85ml"),
    (96,  "VN-GD-3016452", "8934681960075", "BIORE CHARCOAL NOSE CLEANSING STRIP 4P/7",                                "96.BIORE - Miếng Dán Mũi Lột Mụn Biore Than Hoạt Tính 4 Miếng"),
    (97,  "VN-GD-3023693", "8936217700018", "COCOON DAK LAK COFFEE FACE POLISH 150ML",                                 "97.COCOON - Tẩy Tế Bào Cocoon Dak Lak Coffee Làm Sạch Mềm Mại Da Mặt Với Hạt Cà Phê Và Bơ Ca Cao 150ml"),
    (98,  "VN-GD-3010945", "8936032232015", "CLEO HAIR REMOVAL CREAM PINK-SENSITIVE 5",                                "98.CLEO - Kem Tẩy Lông Cléo Dành Cho Da Nhạy Cảm 50g"),
    (99,  "VN-GD-3010943", "8936032232008", "CLEO HAIR REMOVAL CREAM PINK-SENSITIVE25",                                "99.CLEO - Kem Tẩy Lông Cléo Dành Cho Da Nhạy Cảm 25g"),
    (100, "VN-GD-3013352", "8935006537873", "SUNPLAY SUN MILK SUPER BLOCK SPF81 70G",                                  "100.SUNPLAY - Sữa Chống Nắng Cực Mạnh Sunplay Super Block SPF50+ PA++++ 70g"),
    (101, "VN-GD-3025301", "4909978147105", "ANESSA PERFECT UV SUNSCREEN SKINCARE MILK N 60ML",                        "101.ANESSA - Sữa Chống Nắng Anessa Dưỡng Da Kiềm Dầu Bảo Vệ Hoàn Hảo Perfect UV SPF50+ PA++++ 60ml"),
    (102, "VN-GD-3017363", "4909978131586", "ANESSA PERFECT UV MILD MILK SPF50+ 60ML",                                 "102.ANESSA - Sữa Chống Nắng Anessa Cho Da Nhạy Cảm Perfect UV Mild Milk SPF50+/PA++++ 60ml"),
    (103, "VN-GD-3027053", "8994993018086", "LOREAL UV DEFENDER - FLUID SPF50+ 50ML",                                  "103.L'ORÉAL - Dưỡng Chất Chống Nắng L'Oreal Paris Uv Defender Serum Protector Invisible Fluid 50ml"),
    (104, "VN-GD-3024648", "8809576261301", "SKIN1004 CENTELLA PLUS UV SPF50 50ML",                                    "104.SKIN1004 - Kem Chống Nắng Skin1004 Madagascar Centella Air-Fit SunCream Plus SPF50+ PA++++ 50ml"),
    (105, "VN-GD-3029724", "9557514034065", "GUARDIAN GOAT'S MILK MOISTURISING SHOWER CREAM 1L/ 12",                  "105.WATSONS - Kem Tắm Watsons Love My Skin Hương Oải Hương Cream Body Wash 1000ml"),
    (106, "VN-GD-3024536", "9557514021256", "GUARDIAN FACIAL SQUARES 100S/64",                                         "106.WATSONS - Bông Tẩy Trang Watsons Cotton Pads Shibainc Comics 100pcs"),
    (107, "VN-GD-3024541", "9557514021409", "GUARDIAN ULTR SUPER FACIAL COTTON300S/96",                                "107.WATSONS - Bông Tẩy Trang Watsons Square Puffs 150S+40%"),
]
# Quick lookup: barcode → master entry
_MASTER_BY_BC: dict[str, tuple] = {m[2]: m for m in _MASTER_SKUS}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_price(v: Any) -> float | None:
    if v is None:
        return None
    s = str(v).replace(',', '').strip()
    try:
        return float(s)
    except ValueError:
        return None


def _to_time(v: Any):
    """Parse time value → datetime.time (fallback 00:00:00)."""
    from datetime import time as dtime
    if v is None:
        return dtime(0, 0, 0)
    if isinstance(v, dtime):
        return v
    if isinstance(v, datetime):
        return v.time()
    try:
        parts = str(v).strip().split(':')
        return dtime(int(parts[0]), int(parts[1]), int(parts[2] if len(parts) > 2 else 0))
    except Exception:
        return dtime(0, 0, 0)


def _to_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ('%Y-%m-%d', '%B %d, %Y', '%b %d, %Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            pass
    return None


def _extract_stt(product_name: Any) -> int | None:
    """Extract leading number from '01.SOME BY MI...' → 1."""
    if not product_name:
        return None
    m = re.match(r'^(\d+)\.', str(product_name).strip())
    return int(m.group(1)) if m else None


def _iso_week(d: date) -> int:
    # Week runs Fri→Thu: shift back 4 days so Fri aligns to Mon in ISO calendar
    from datetime import timedelta
    return (d - timedelta(days=4)).isocalendar()[1]


def _parse_product_code(code: Any) -> tuple[str | None, str | None]:
    """'01-3021691-8809647390015' → ('VN-GD-3021691', '8809647390015')"""
    if not code:
        return None, None
    parts = str(code).split('-')
    if len(parts) >= 3:
        return f"VN-GD-{parts[1]}", parts[2]
    return None, None


def _others_remarks(row: dict) -> str | None:
    if row.get('mua_tang') == 'Yes':
        q1, q2 = row.get('qty1'), row.get('qty2')
        if q1 is not None and q2 is not None:
            return f"Buy {int(q1)} Get {int(q2)}"
    if row.get('mua_gia') == 'Yes':
        q3, g2 = row.get('qty3'), _parse_price(row.get('gia2'))
        if q3 is not None and g2 is not None:
            return f"Buy {int(q3)} Pay {int(g2):,}"
    return None


def _compute_prices(row: dict) -> tuple[float | None, float | None, float | None]:
    """Returns (price, regular, promo_off)."""
    gia_ban = _parse_price(row.get('gia_ban'))
    gia_goc = _parse_price(row.get('gia_goc'))

    if gia_ban is None:
        return None, None, None

    if row.get('giam') == 'Yes' and gia_goc is not None:
        # Case 1: normal discount
        promo_off = (gia_goc - gia_ban) / gia_goc if gia_goc > 0 else 0
        return gia_ban, gia_goc, promo_off

    if row.get('mua_tang') == 'Yes':
        # Case 3: Buy X Get Y → Regular = Giá bán, Price = Regular × Qty1 / (Qty1+Qty2)
        q1, q2 = row.get('qty1'), row.get('qty2')
        if q1 is not None and q2 is not None and (q1 + q2) > 0:
            price = gia_ban * q1 / (q1 + q2)
            promo_off = (gia_ban - price) / gia_ban if gia_ban > 0 else 0
            return price, gia_ban, promo_off
        return gia_ban, gia_ban, 0

    if row.get('mua_gia') == 'Yes':
        # Case 4: Buy X for Y → Price = Giá2 / Qty3, Regular = Giá bán
        q3, g2 = row.get('qty3'), _parse_price(row.get('gia2'))
        if q3 is not None and q3 > 0 and g2 is not None:
            price = g2 / q3
            promo_off = (gia_ban - price) / gia_ban if gia_ban > 0 else 0
            return price, gia_ban, promo_off
        return gia_ban, gia_ban, 0

    # Case 2: no promo
    return gia_ban, gia_ban, 0


# ── Parse ─────────────────────────────────────────────────────────────────────

def parse_guardian(xlsx_bytes: bytes) -> list[dict]:
    """Parse Guardian.xlsx → list of row dicts (only Bán=Yes rows)."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb['Q1']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Row 4 (index 3) = header, data from index 4
    header = rows[3]
    col = {str(h).strip(): i for i, h in enumerate(header) if h}

    def g(row: tuple, name: str) -> Any:
        idx = col.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    records = []
    for row in rows[4:]:
        if row[0] is None:
            continue

        d = _to_date(g(row, 'Date'))
        if d is None:
            continue

        article, barcode = _parse_product_code(g(row, 'Product code'))
        if not barcode:
            continue

        ban = g(row, 'Sản phẩm có bán không?')

        records.append({
            'date':         d,
            'time':         _to_time(g(row, 'Time')),
            'week':         _iso_week(d),
            'article':      article,
            'barcode':      barcode,
            'description':  g(row, 'Description'),
            'product_name': g(row, 'Product name'),
            'ban':          ban,                       # 'Yes' / 'No' / None
            'gia_ban':      g(row, 'Giá bán')      if ban == 'Yes' else None,
            'gia_goc':      g(row, 'Giá gốc')      if ban == 'Yes' else None,
            'giam':         g(row, 'Sản phẩm có giảm giá không?') if ban == 'Yes' else None,
            'promo_start':  _to_date(g(row, 'Từ ngày'))            if ban == 'Yes' else None,
            'promo_end':    _to_date(g(row, 'Đến ngày'))           if ban == 'Yes' else None,
            'mua_tang':     g(row, 'Có [Mua ... tặng ... không?]') if ban == 'Yes' else None,
            'qty1':         g(row, 'Mua với số lượng ...')         if ban == 'Yes' else None,
            'qty2':         g(row, 'Thì tặng với số lượng ...')    if ban == 'Yes' else None,
            'mua_gia':      g(row, 'Có [Mua ... với giá ... không?]') if ban == 'Yes' else None,
            'qty3':         g(row, 'Mua với số lượng là')          if ban == 'Yes' else None,
            'gia2':         g(row, 'Thì có giá là')                if ban == 'Yes' else None,
            'online':       g(row, 'Giá được lấy từ online?')      if ban == 'Yes' else None,
            'trang_thai':   g(row, 'Trạng thái không bán?')        if ban != 'Yes' else None,
        })

    return records


# ── Aggregate ─────────────────────────────────────────────────────────────────

def aggregate_skus(
    records: list[dict],
    curr_week: int | None = None,
    prev_week: int | None = None,
) -> tuple[list[dict], int, int, list[int]]:
    """Group by (barcode, week), pick latest (date, time) row per SKU.
    Returns (sku_list, curr_week, prev_week, available_weeks).
    """
    by_bc_week: dict[tuple, list] = defaultdict(list)
    for r in records:
        by_bc_week[(r['barcode'], r['week'])].append(r)

    available_weeks = sorted({w for _, w in by_bc_week})
    if curr_week is None:
        curr_week = max(available_weeks)
    if prev_week is None:
        prev_week = curr_week - 1

    def _best(rows: list[dict]) -> dict | None:
        """Pick the single latest row by (date, time) — most recent survey entry wins."""
        if not rows:
            return None
        return max(rows, key=lambda r: (r['date'], r.get('time')))

    results = []
    for (stt, article, barcode, description, product_name) in _MASTER_SKUS:
        bc = barcode
        curr_rows = by_bc_week.get((bc, curr_week), [])
        prev_rows = by_bc_week.get((bc, prev_week), [])

        curr = _best(curr_rows)
        prev = _best(prev_rows)

        cp, cr, co = _compute_prices(curr) if curr else (None, None, None)
        pp, pr, po = _compute_prices(prev) if prev else (None, None, None)

        remarks = _others_remarks(curr) if curr else (_others_remarks(prev) if prev else None)

        curr_sold = curr is not None and curr.get('ban') == 'Yes'
        prev_sold = prev is not None and prev.get('ban') == 'Yes'

        # Price column: no data → "—" (None); not-sold → trang_thai or "OOS"
        curr_status = None if not curr_rows else (((curr.get('trang_thai') if curr else None) or 'OOS') if not curr_sold else None)
        prev_status = ((prev.get('trang_thai') if prev else None) or 'OOS') if not prev_sold else None

        # Agency comment: "Online" only when the week's sold row sourced price online (OOS excluded)
        is_online = (curr_sold and curr.get('online') == 'Yes') or (prev_sold and prev.get('online') == 'Yes')
        agency_comment: str | None = 'Online' if is_online else None

        # Error: no data collected yet for curr_week → flag immediately
        error: str | None = None
        if not curr_rows:
            error = f"Chưa có data W{curr_week}"
        elif curr_sold and prev_sold:
            parts = []
            cp_i = int(round(cp)) if cp is not None else None
            pp_i = int(round(pp)) if pp is not None else None
            cr_i = int(round(cr)) if cr is not None else None
            pr_i = int(round(pr)) if pr is not None else None
            if cp_i is not None and pp_i is not None and cp_i != pp_i:
                pct = abs(cp_i - pp_i) / pp_i * 100
                dir_ = 'TĂNG' if cp_i > pp_i else 'GIẢM'
                parts.append(f"Giá bán: W{curr_week}={cp_i:,} {dir_} {pct:.2f}% so với W{prev_week}={pp_i:,}")
            if cr_i is not None and pr_i is not None and cr_i != pr_i:
                pct = abs(cr_i - pr_i) / pr_i * 100
                dir_ = 'TĂNG' if cr_i > pr_i else 'GIẢM'
                parts.append(f"Giá gốc: W{curr_week}={cr_i:,} {dir_} {pct:.2f}% so với W{prev_week}={pr_i:,}")
            error = ' | '.join(parts) if parts else None

        results.append({
            'stt':            stt,
            'article':        article,
            'barcode':        barcode,
            'description':    description,
            'product_name':   product_name,
            'curr_price':     cp,
            'curr_regular':   cr,
            'curr_promo_off': co,
            'curr_status':    curr_status,
            'promo_start':    curr['promo_start'] if curr else None,
            'promo_end':      curr['promo_end'] if curr else None,
            'remarks':        remarks,
            'agency_comment': agency_comment,
            'prev_price':     pp,
            'prev_regular':   pr,
            'prev_promo_off': po,
            'prev_status':    prev_status,
            'error':          error,
        })

    results.sort(key=lambda s: s['stt'])
    return results, curr_week, prev_week, available_weeks


# ── Generate xlsx ─────────────────────────────────────────────────────────────

def _thin() -> Border:
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def generate_xlsx(skus: list[dict], curr_week: int, prev_week: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Watsons"

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    ctr      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border   = _thin()

    headers = [
        'STT', 'Aricle Number (banner)', 'BARCODE',
        'Product Description', 'Product Description Detail',
        f'Price W{curr_week}', f'Regular price W{curr_week}', f'Promo OFF W{curr_week}',
        'Promo Start Date', 'Promo End Date', 'Others Remarks', "Agency's comment",
        f'Price W{prev_week}', f'Regular price W{prev_week}', f'Promo OFF W{prev_week}',
        'ERRORS',
    ]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = ctr
        c.border = border

    alt_fill   = PatternFill("solid", fgColor="DEEAF1")
    err_fill   = PatternFill("solid", fgColor="FCE4D6")
    norm_font  = Font(size=10)
    num_fmt    = '#,##0'
    pct_fmt    = '0.00%'
    date_fmt   = 'D-MMM-YY'

    for ri, sku in enumerate(skus, 2):
        fill = err_fill if sku['error'] else (alt_fill if ri % 2 == 0 else None)
        curr_price_val = sku.get('curr_status') or sku['curr_price']
        prev_price_val = sku.get('prev_status') or sku['prev_price']
        vals = [
            sku['stt'], sku['article'], sku['barcode'],
            sku['description'], sku['product_name'],
            curr_price_val, sku['curr_regular'], sku['curr_promo_off'],
            sku['promo_start'], sku['promo_end'], sku['remarks'], sku.get('agency_comment'),
            prev_price_val, sku['prev_regular'], sku['prev_promo_off'],
            sku['error'],
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = norm_font
            c.border = border
            c.alignment = Alignment(vertical='center')
            if fill:
                c.fill = fill
            h = headers[ci - 1]
            if ('Price' in h or 'Regular' in h) and not isinstance(val, str):
                c.number_format = num_fmt
            elif 'Promo OFF' in h:
                c.number_format = pct_fmt
            elif 'Date' in h and isinstance(val, date):
                c.number_format = date_fmt

    col_widths = [5, 18, 16, 42, 58, 14, 14, 12, 14, 14, 18, 16, 14, 14, 12, 30]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = w

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Entry point ───────────────────────────────────────────────────────────────

def _fmt_date(d: Any) -> str | None:
    if d is None:
        return None
    if hasattr(d, 'strftime'):
        return d.strftime('%d-%b-%Y')
    return str(d)


def _fmt_price(v: Any) -> str | None:
    if v is None:
        return None
    try:
        return f"{int(round(v)):,}"
    except (TypeError, ValueError):
        return None


def _fmt_pct(v: Any) -> str | None:
    if v is None:
        return None
    try:
        return f"{v * 100:.2f}%"
    except (TypeError, ValueError):
        return None


def process(
    xlsx_bytes: bytes,
    curr_week: int | None = None,
    prev_week: int | None = None,
) -> dict:
    records = parse_guardian(xlsx_bytes)
    if not records:
        raise ValueError("No valid rows found — check that the file has a 'Q1' sheet with Bán=Yes rows")

    skus, curr_week, prev_week, available_weeks = aggregate_skus(records, curr_week, prev_week)
    n_errors = sum(1 for s in skus if s['error'])

    xlsx_out = generate_xlsx(skus, curr_week, prev_week)

    return {
        'filename':        f"DFI_PriceCheck_W{curr_week}_vs_W{prev_week}.xlsx",
        'week_current':    f'W{curr_week}',
        'week_prev':       f'W{prev_week}',
        'available_weeks': [f'W{w}' for w in available_weeks],
        'n_skus':       len(skus),
        'n_errors':     n_errors,
        'base64':       base64.b64encode(xlsx_out).decode('ascii'),
        'skus': [
            {
                'stt':            s['stt'],
                'product_name':   s['product_name'],
                'curr_price':     s.get('curr_status') or _fmt_price(s['curr_price']),
                'curr_regular':   _fmt_price(s['curr_regular']),
                'curr_promo_off': _fmt_pct(s['curr_promo_off']),
                'promo_start':    _fmt_date(s['promo_start']),
                'promo_end':      _fmt_date(s['promo_end']),
                'remarks':        s['remarks'],
                'agency_comment': s.get('agency_comment'),
                'prev_price':     s.get('prev_status') or _fmt_price(s['prev_price']),
                'prev_regular':   _fmt_price(s['prev_regular']),
                'prev_promo_off': _fmt_pct(s['prev_promo_off']),
                'error':          s['error'],
            }
            for s in skus
        ],
    }
